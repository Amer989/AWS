
import boto3
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

def get_aws_daily_costs(start_date, end_date):
    """
    Fetch daily AWS cost data using Cost Explorer API
    """
    try:
        # Initialize the Cost Explorer client
        ce_client = boto3.client('ce')
        
        # Request daily cost data from AWS Cost Explorer
        response = ce_client.get_cost_and_usage(
            TimePeriod={
                'Start': start_date,
                'End': end_date
            },
            Granularity='DAILY',
            Metrics=['UnblendedCost'],
            GroupBy=[
                {
                    'Type': 'DIMENSION',
                    'Key': 'SERVICE'
                }
            ]
        )
        
        return response
    
    except Exception as e:
        print(f"Error fetching AWS cost data: {str(e)}")
        return None

def parse_cost_data(cost_response):
    """
    Parse the AWS Cost Explorer response into a structured format
    """
    if not cost_response:
        return None
    
    results = []
    
    for time_period in cost_response['ResultsByTime']:
        date = time_period['TimePeriod']['Start']
        
        for group in time_period['Groups']:
            service = group['Keys'][0]
            amount = float(group['Metrics']['UnblendedCost']['Amount'])
            currency = group['Metrics']['UnblendedCost']['Unit']
            
            results.append({
                'Date': date,
                'Service': service,
                'Cost': amount,
                'Currency': currency
            })
    
    # Add the total cost for each day
    for time_period in cost_response['ResultsByTime']:
        date = time_period['TimePeriod']['Start']
        if 'Total' in time_period:
            total_amount = float(time_period['Total']['UnblendedCost']['Amount'])
            currency = time_period['Total']['UnblendedCost']['Unit']
            
            results.append({
                'Date': date,
                'Service': 'Total',
                'Cost': total_amount,
                'Currency': currency
            })
    
    return results

def create_excel_report(cost_data, output_file):
    """
    Create an Excel report with the AWS cost data
    """
    if not cost_data:
        print("No cost data available to create the report.")
        return False
    
    # Convert to DataFrame
    df = pd.DataFrame(cost_data)
    
    # Create a pivot table with services as columns and dates as rows
    pivot_df = df.pivot_table(
        index='Date', 
        columns='Service', 
        values='Cost', 
        aggfunc='sum'
    ).reset_index()
    
    # Format the date column
    pivot_df['Date'] = pd.to_datetime(pivot_df['Date']).dt.strftime('%Y-%m-%d')
    
    # Create the Excel writer object
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Write the DataFrame to Excel
    pivot_df.to_excel(writer, sheet_name='Daily Costs', index=False)
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Daily Costs']
    
    # Apply formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    for col_num, column_title in enumerate(pivot_df.columns):
        cell = worksheet.cell(row=1, column=col_num+1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Set column width
        worksheet.column_dimensions[get_column_letter(col_num+1)].width = 15
    
    # Format all cells with borders and number format for cost columns
    for row_num in range(2, len(pivot_df) + 2):
        for col_num in range(1, len(pivot_df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Apply number format to cost columns (skip the Date column)
            if col_num > 1:
                cell.number_format = '$#,##0.00'
    
    # Create a bar chart for total costs by day
    if 'Total' in pivot_df.columns:
        chart_sheet = workbook.create_sheet('Cost Chart')
        
        chart = BarChart()
        chart.title = "Daily AWS Costs"
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Cost (USD)"
        
        # Get the column index for 'Total'
        total_col_idx = list(pivot_df.columns).index('Total') + 1
        
        data = Reference(worksheet, min_col=total_col_idx, max_col=total_col_idx, 
                        min_row=1, max_row=len(pivot_df) + 1)
        cats = Reference(worksheet, min_col=1, max_col=1, 
                        min_row=2, max_row=len(pivot_df) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 15
        chart.width = 25
        
        chart_sheet.add_chart(chart, "B3")
    
    # Create summary sheet
    summary_sheet = workbook.create_sheet('Cost Summary')
    
    # Get total cost by service
    service_totals = df[df['Service'] != 'Total'].groupby('Service')['Cost'].sum().reset_index()
    service_totals = service_totals.sort_values('Cost', ascending=False)
    
    # Write service totals to summary sheet
    summary_sheet['A1'] = 'Service'
    summary_sheet['B1'] = 'Total Cost'
    summary_sheet['C1'] = 'Percentage'
    
    # Apply formatting to headers
    for col in ['A1', 'B1', 'C1']:
        summary_sheet[col].fill = header_fill
        summary_sheet[col].font = header_font
        summary_sheet[col].alignment = Alignment(horizontal='center')
        summary_sheet[col].border = border
    
    # Set column width
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 15
    
    # Calculate total cost for percentage calculation
    total_cost = service_totals['Cost'].sum()
    
    # Write service data
    for idx, row in enumerate(service_totals.itertuples(), start=2):
        summary_sheet[f'A{idx}'] = row.Service
        summary_sheet[f'B{idx}'] = row.Cost
        summary_sheet[f'B{idx}'].number_format = '$#,##0.00'
        
        # Calculate and format percentage
        percentage = (row.Cost / total_cost) * 100 if total_cost > 0 else 0
        summary_sheet[f'C{idx}'] = percentage / 100  # Excel uses decimal for percentage format
        summary_sheet[f'C{idx}'].number_format = '0.00%'
        
        # Apply borders
        for col in ['A', 'B', 'C']:
            summary_sheet[f'{col}{idx}'].border = border
    
    # Add total row
    total_row = len(service_totals) + 2
    summary_sheet[f'A{total_row}'] = 'Total'
    summary_sheet[f'A{total_row}'].font = Font(bold=True)
    summary_sheet[f'B{total_row}'] = total_cost
    summary_sheet[f'B{total_row}'].number_format = '$#,##0.00'
    summary_sheet[f'B{total_row}'].font = Font(bold=True)
    
    # Create a pie chart for service distribution
    chart = BarChart()
    chart.type = 'col'
    chart.title = "Cost Distribution by Service"
    chart.x_axis.title = "Service"
    chart.y_axis.title = "Cost (USD)"
    
    data = Reference(summary_sheet, min_col=2, max_col=2, 
                     min_row=1, max_row=min(11, len(service_totals) + 1))  # Top 10 services
    cats = Reference(summary_sheet, min_col=1, max_col=1, 
                     min_row=2, max_row=min(11, len(service_totals) + 1))
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 15
    chart.width = 25
    
    summary_sheet.add_chart(chart, "E3")
    
    # Save the workbook
    writer.close()
    
    return True

def main():
    # Define the date range for the report (last 30 days)
    end_date = datetime.now().strftime('%Y-%m-%d')
    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    print(f"Fetching AWS cost data from {start_date} to {end_date}...")
    
    # Get cost data
    cost_response = get_aws_daily_costs(start_date, end_date)
    
    if cost_response:
        # Parse cost data
        cost_data = parse_cost_data(cost_response)
        
        # Create Excel report
        output_file = f"AWS_Daily_Costs_{datetime.now().strftime('%Y%m%d')}.xlsx"
        if create_excel_report(cost_data, output_file):
            print(f"Excel report created successfully: {output_file}")
        else:
            print("Failed to create Excel report.")
    else:
        print("Failed to fetch AWS cost data.")

if __name__ == "__main__":
    main()
