import boto3
import pandas as pd
import calendar
from datetime import datetime, timedelta
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import argparse

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='Generate monthly AWS cost reports')
    parser.add_argument('--month', type=int, help='Month number (1-12)')
    parser.add_argument('--year', type=int, help='Year (YYYY)')
    parser.add_argument('--output', type=str, help='Output file path', default=None)
    parser.add_argument('--profile', type=str, help='AWS profile name', default=None)
    parser.add_argument('--region', type=str, help='AWS region', default='us-east-1')
    parser.add_argument('--group-by', type=str, choices=['service', 'account', 'region', 'tag'], 
                        default='service', help='Group costs by dimension')
    parser.add_argument('--tag-key', type=str, help='Tag key to group by (if group-by=tag)')
    
    args = parser.parse_args()
    
    # Set defaults for month and year if not provided
    if args.month is None:
        # Default to previous month
        today = datetime.now()
        if today.month == 1:
            args.month = 12
            args.year = today.year - 1
        else:
            args.month = today.month - 1
            args.year = today.year
    
    if args.year is None:
        args.year = datetime.now().year
    
    # Validate month
    if args.month < 1 or args.month > 12:
        raise ValueError("Month must be between 1 and 12")
    
    # Set default output file if not provided
    if args.output is None:
        month_name = calendar.month_name[args.month]
        args.output = f"AWS_Monthly_Costs_{args.year}_{month_name}.xlsx"
    
    # Additional validation for tag-based grouping
    if args.group_by == 'tag' and args.tag_key is None:
        raise ValueError("--tag-key is required when --group-by=tag")
    
    return args

def get_aws_monthly_costs(year, month, group_by='service', tag_key=None, profile=None, region='us-east-1'):
    """Fetch monthly AWS cost data using Cost Explorer API"""
    try:
        # Calculate start and end dates for the month
        start_date = f"{year}-{month:02d}-01"
        
        # Calculate the last day of the month
        _, last_day = calendar.monthrange(year, month)
        end_date = f"{year}-{month:02d}-{last_day + 1:02d}"  # CE API end date is exclusive
        
        # Create boto3 session with optional profile
        if profile:
            session = boto3.Session(profile_name=profile, region_name=region)
        else:
            session = boto3.Session(region_name=region)
        
        # Initialize the Cost Explorer client
        ce_client = session.client('ce')
        
        # Determine GroupBy configuration based on input
        if group_by == 'service':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'SERVICE'}]
        elif group_by == 'account':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'LINKED_ACCOUNT'}]
        elif group_by == 'region':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'REGION'}]
        elif group_by == 'tag' and tag_key:
            group_by_config = [{'Type': 'TAG', 'Key': tag_key}]
        else:
            raise ValueError(f"Invalid group_by: {group_by}")
        
        # Request cost data from AWS Cost Explorer
        response = ce_client.get_cost_and_usage(
            TimePeriod={
                'Start': start_date,
                'End': end_date
            },
            Granularity='DAILY',
            Metrics=['UnblendedCost', 'UsageQuantity'],
            GroupBy=group_by_config
        )
        
        return response
    
    except Exception as e:
        print(f"Error fetching AWS cost data: {str(e)}")
        return None

def parse_cost_data(cost_response, group_by='service'):
    """Parse the AWS Cost Explorer response into a structured format"""
    if not cost_response:
        return None
    
    results = []
    
    # Determine the name of the dimension based on group_by
    dimension_name = group_by.capitalize()
    if group_by == 'tag':
        dimension_name = 'Tag'
    
    for time_period in cost_response['ResultsByTime']:
        date = time_period['TimePeriod']['Start']
        
        for group in time_period['Groups']:
            dimension_value = group['Keys'][0]
            
            # For services, strip the prefix
            if group_by == 'service' and dimension_value.startswith('Amazon'):
                dimension_value = dimension_value.replace('Amazon ', '')
            
            amount = float(group['Metrics']['UnblendedCost']['Amount'])
            currency = group['Metrics']['UnblendedCost']['Unit']
            usage = float(group['Metrics']['UsageQuantity']['Amount'])
            
            results.append({
                'Date': date,
                dimension_name: dimension_value,
                'Cost': amount,
                'Currency': currency,
                'Usage': usage
            })
    
    # Add the total cost for each day
    for time_period in cost_response['ResultsByTime']:
        date = time_period['TimePeriod']['Start']
        if 'Total' in time_period:
            total_amount = float(time_period['Total']['UnblendedCost']['Amount'])
            currency = time_period['Total']['UnblendedCost']['Unit']
            total_usage = float(time_period['Total']['UsageQuantity']['Amount'])
            
            results.append({
                'Date': date,
                dimension_name: 'Total',
                'Cost': total_amount,
                'Currency': currency,
                'Usage': total_usage
            })
    
    return results

def create_excel_report(cost_data, output_file, month, year, group_by='service'):
    """Create an Excel report with the AWS cost data"""
    if not cost_data:
        print("No cost data available to create the report.")
        return False
    
    # Convert to DataFrame and ensure the dimension column has the right name
    df = pd.DataFrame(cost_data)
    dimension_name = group_by.capitalize()
    if group_by == 'tag':
        dimension_name = 'Tag'
    
    # Create a pivot table with services as columns and dates as rows
    pivot_df = df.pivot_table(
        index='Date', 
        columns=dimension_name, 
        values='Cost', 
        aggfunc='sum'
    ).reset_index()
    
    # Format the date column
    pivot_df['Date'] = pd.to_datetime(pivot_df['Date']).dt.strftime('%Y-%m-%d')
    
    # Calculate month-to-date total by dimension
    monthly_totals = df.groupby(dimension_name)['Cost'].sum().reset_index()
    monthly_totals = monthly_totals.sort_values('Cost', ascending=False)
    
    # Create the Excel writer object
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Write the daily breakdown DataFrame to Excel
    pivot_df.to_excel(writer, sheet_name='Daily Breakdown', index=False)
    
    # Write the monthly totals to Excel
    monthly_totals.to_excel(writer, sheet_name='Monthly Summary', index=False)
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    daily_sheet = writer.sheets['Daily Breakdown']
    summary_sheet = writer.sheets['Monthly Summary']
    
    # Apply formatting (common style elements)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format Daily Breakdown sheet
    # Format headers
    for col_num, column_title in enumerate(pivot_df.columns):
        cell = daily_sheet.cell(row=1, column=col_num+1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Set column width
        daily_sheet.column_dimensions[get_column_letter(col_num+1)].width = 15
    
    # Format all cells with borders and number format for cost columns
    for row_num in range(2, len(pivot_df) + 2):
        for col_num in range(1, len(pivot_df.columns) + 1):
            cell = daily_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Apply number format to cost columns (skip the Date column)
            if col_num > 1:
                cell.number_format = '$#,##0.00'
    
    # Add title and report info to the daily sheet
    daily_sheet.insert_rows(0, 2)
    month_name = calendar.month_name[month]
    daily_sheet['A1'] = f"AWS Daily Cost Breakdown - {month_name} {year}"
    daily_sheet['A1'].font = Font(size=14, bold=True)
    daily_sheet.merge_cells('A1:E1')
    
    # Format Monthly Summary sheet
    # Format headers
    for col_num, column_title in enumerate(monthly_totals.columns):
        cell = summary_sheet.cell(row=1, column=col_num+1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Set column width
        summary_sheet.column_dimensions[get_column_letter(col_num+1)].width = 20
    
    # Format data cells
    for row_num in range(2, len(monthly_totals) + 2):
        for col_num in range(1, len(monthly_totals.columns) + 1):
            cell = summary_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Apply number format to cost column
            if col_num == 2:  # Cost column
                cell.number_format = '$#,##0.00'
    
    # Add percentage column to summary
    total_cost = monthly_totals['Cost'].sum()
    summary_sheet['C1'] = 'Percentage'
    summary_sheet['C1'].fill = header_fill
    summary_sheet['C1'].font = header_font
    summary_sheet['C1'].alignment = Alignment(horizontal='center')
    summary_sheet['C1'].border = border
    
    for row_num in range(2, len(monthly_totals) + 2):
        cost_cell = summary_sheet.cell(row=row_num, column=2)
        percentage = float(cost_cell.value) / total_cost if total_cost > 0 else 0
        
        cell = summary_sheet.cell(row=row_num, column=3)
        cell.value = percentage
        cell.number_format = '0.00%'
        cell.border = border
    
    # Add title and report info to the summary sheet
    summary_sheet.insert_rows(0, 2)
    summary_sheet['A1'] = f"AWS Monthly Cost Summary - {month_name} {year}"
    summary_sheet['A1'].font = Font(size=14, bold=True)
    summary_sheet.merge_cells('A1:C1')
    
    # Add total row to summary
    total_row = len(monthly_totals) + 3  # +3 because we added 2 title rows
    summary_sheet[f'A{total_row}'] = 'Total'
    summary_sheet[f'A{total_row}'].font = Font(bold=True)
    summary_sheet[f'B{total_row}'] = total_cost
    summary_sheet[f'B{total_row}'].number_format = '$#,##0.00'
    summary_sheet[f'B{total_row}'].font = Font(bold=True)
    summary_sheet[f'C{total_row}'] = 1
    summary_sheet[f'C{total_row}'].number_format = '0.00%'
    summary_sheet[f'C{total_row}'].font = Font(bold=True)
    
    # Create a pie chart for cost distribution
    pie = PieChart()
    pie.title = f"Cost Distribution by {dimension_name} - {month_name} {year}"
    
    # Get data for top 10 items (to avoid cluttering the chart)
    max_items = min(10, len(monthly_totals))
    data = Reference(summary_sheet, min_col=2, max_col=2, 
                     min_row=3, max_row=2+max_items)
    cats = Reference(summary_sheet, min_col=1, max_col=1, 
                     min_row=3, max_row=2+max_items)
    
    pie.add_data(data)
    pie.set_categories(cats)
    pie.height = 15
    pie.width = 10
    
    # Add the chart to the summary sheet
    summary_sheet.add_chart(pie, "E5")
    
    # Create a visual daily cost chart
    chart_sheet = workbook.create_sheet('Daily Cost Chart')
    
    chart = BarChart()
    chart.title = f"Daily AWS Costs - {month_name} {year}"
    chart.x_axis.title = "Date"
    chart.y_axis.title = "Cost (USD)"
    
    # Get the column index for 'Total' if it exists
    if 'Total' in pivot_df.columns:
        total_col_idx = list(pivot_df.columns).index('Total') + 1
        
        # Add 2 to row references because of the title rows we inserted
        data = Reference(daily_sheet, min_col=total_col_idx, max_col=total_col_idx, 
                        min_row=3, max_row=len(pivot_df) + 3)
        cats = Reference(daily_sheet, min_col=1, max_col=1, 
                        min_row=4, max_row=len(pivot_df) + 3)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 15
        chart.width = 25
        
        chart_sheet.add_chart(chart, "B5")
    
    # Add title to chart sheet
    chart_sheet['A1'] = f"AWS Daily Cost Visualization - {month_name} {year}"
    chart_sheet['A1'].font = Font(size=14, bold=True)
    
    # Create a trends sheet for tracking month-over-month changes
    trends_sheet = workbook.create_sheet('Cost Trends')
    trends_sheet['A1'] = "Month-over-Month Cost Trends"
    trends_sheet['A1'].font = Font(size=14, bold=True)
    trends_sheet['A3'] = "To track cost trends over time, run this report for multiple months and consolidate the data here."
    
    # Add report metadata
    info_sheet = workbook.create_sheet('Report Info')
    info_sheet['A1'] = "AWS Cost Report - Metadata"
    info_sheet['A1'].font = Font(size=14, bold=True)
    
    info_sheet['A3'] = "Report Generated:"
    info_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    info_sheet['A4'] = "Report Period:"
    info_sheet['B4'] = f"{month_name} {year}"
    
    info_sheet['A5'] = "Grouping Dimension:"
    info_sheet['B5'] = dimension_name
    
    info_sheet['A6'] = "Total Monthly Cost:"
    info_sheet['B6'] = total_cost
    info_sheet['B6'].number_format = '$#,##0.00'
    
    info_sheet['A8'] = "Notes:"
    info_sheet['A9'] = "- Cost data is based on AWS Cost Explorer API and may have a delay of 24-48 hours"
    info_sheet['A10'] = "- The report includes unblended costs"
    info_sheet['A11'] = f"- Costs are grouped by {dimension_name.lower()}"
    
    # Save the workbook
    writer.close()
    
    return True

def main():
    # Parse command line arguments
    args = parse_arguments()
    
    month_name = calendar.month_name[args.month]
    print(f"Generating AWS cost report for {month_name} {args.year}...")
    
    # Get cost data
    cost_response = get_aws_monthly_costs(
        args.year, 
        args.month, 
        group_by=args.group_by,
        tag_key=args.tag_key,
        profile=args.profile,
        region=args.region
    )
    
    if cost_response:
        # Parse cost data
        cost_data = parse_cost_data(cost_response, args.group_by)
        
        # Create Excel report
        if create_excel_report(cost_data, args.output, args.month, args.year, args.group_by):
            print(f"Excel report created successfully: {args.output}")
        else:
            print("Failed to create Excel report.")
    else:
        print("Failed to fetch AWS cost data.")

if __name__ == "__main__":
    main()
