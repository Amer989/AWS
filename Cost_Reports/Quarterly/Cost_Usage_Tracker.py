import boto3
import pandas as pd
import calendar
from datetime import datetime, timedelta
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import argparse

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='Generate quarterly AWS cost reports')
    parser.add_argument('--quarter', type=int, help='Quarter number (1-4)')
    parser.add_argument('--year', type=int, help='Year (YYYY)')
    parser.add_argument('--output', type=str, help='Output file path', default=None)
    parser.add_argument('--profile', type=str, help='AWS profile name', default=None)
    parser.add_argument('--region', type=str, help='AWS region', default='us-east-1')
    parser.add_argument('--group-by', type=str, choices=['service', 'account', 'region', 'tag'], 
                        default='service', help='Group costs by dimension')
    parser.add_argument('--tag-key', type=str, help='Tag key to group by (if group-by=tag)')
    parser.add_argument('--granularity', type=str, choices=['daily', 'monthly'], 
                        default='monthly', help='Data granularity')
    
    args = parser.parse_args()
    
    # Set defaults for quarter and year if not provided
    if args.quarter is None:
        # Default to current quarter
        today = datetime.now()
        args.quarter = (today.month - 1) // 3 + 1
        
    if args.year is None:
        args.year = datetime.now().year
    
    # Validate quarter
    if args.quarter < 1 or args.quarter > 4:
        raise ValueError("Quarter must be between 1 and 4")
    
    # Set default output file if not provided
    if args.output is None:
        args.output = f"AWS_Quarterly_Costs_Q{args.quarter}_{args.year}.xlsx"
    
    # Additional validation for tag-based grouping
    if args.group_by == 'tag' and args.tag_key is None:
        raise ValueError("--tag-key is required when --group-by=tag")
    
    return args

def get_quarter_dates(year, quarter):
    """Get start and end dates for a given quarter"""
    if quarter == 1:
        start_month = 1
    elif quarter == 2:
        start_month = 4
    elif quarter == 3:
        start_month = 7
    elif quarter == 4:
        start_month = 10
    else:
        raise ValueError("Quarter must be between 1 and 4")
    
    start_date = f"{year}-{start_month:02d}-01"
    
    # End date is the first day of the next quarter
    if quarter < 4:
        end_month = start_month + 3
        end_year = year
    else:
        end_month = 1
        end_year = year + 1
    
    end_date = f"{end_year}-{end_month:02d}-01"
    
    return start_date, end_date

def get_aws_quarterly_costs(year, quarter, group_by='service', tag_key=None, 
                            profile=None, region='us-east-1', granularity='monthly'):
    """Fetch quarterly AWS cost data using Cost Explorer API"""
    try:
        # Get start and end dates for the quarter
        start_date, end_date = get_quarter_dates(year, quarter)
        
        # Create boto3 session with optional profile
        if profile:
            session = boto3.Session(profile_name=profile, region_name=region)
        else:
            session = boto3.Session(region_name=region)
        
        # Initialize the Cost Explorer client
        ce_client = session.client('ce')
        
        # Determine GroupBy configuration based on input
        if group_by == 'service':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'SERVICE'}]
        elif group_by == 'account':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'LINKED_ACCOUNT'}]
        elif group_by == 'region':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'REGION'}]
        elif group_by == 'tag' and tag_key:
            group_by_config = [{'Type': 'TAG', 'Key': tag_key}]
        else:
            raise ValueError(f"Invalid group_by: {group_by}")
        
        # Convert granularity string to API format
        ce_granularity = 'DAILY' if granularity == 'daily' else 'MONTHLY'
        
        # Request cost data from AWS Cost Explorer
        response = ce_client.get_cost_and_usage(
            TimePeriod={
                'Start': start_date,
                'End': end_date
            },
            Granularity=ce_granularity,
            Metrics=['UnblendedCost', 'UsageQuantity'],
            GroupBy=group_by_config
        )
        
        return response
    
    except Exception as e:
        print(f"Error fetching AWS cost data: {str(e)}")
        return None

def parse_cost_data(cost_response, group_by='service'):
    """Parse the AWS Cost Explorer response into a structured format"""
    if not cost_response:
        return None
    
    results = []
    
    # Determine the name of the dimension based on group_by
    dimension_name = group_by.capitalize()
    if group_by == 'tag':
        dimension_name = 'Tag'
    
    for time_period in cost_response['ResultsByTime']:
        period_start = time_period['TimePeriod']['Start']
        period_end = time_period['TimePeriod']['End']
        
        # For monthly granularity, use the month name
        date_obj = datetime.strptime(period_start, '%Y-%m-%d')
        if period_end > period_start:
            # This is for monthly granularity
            period_label = date_obj.strftime('%B %Y')
        else:
            # This is for daily granularity
            period_label = period_start
        
        for group in time_period['Groups']:
            dimension_value = group['Keys'][0]
            
            # For services, strip the prefix
            if group_by == 'service' and dimension_value.startswith('Amazon'):
                dimension_value = dimension_value.replace('Amazon ', '')
            
            amount = float(group['Metrics']['UnblendedCost']['Amount'])
            currency = group['Metrics']['UnblendedCost']['Unit']
            usage = float(group['Metrics']['UsageQuantity']['Amount'])
            
            results.append({
                'Period': period_label,
                'StartDate': period_start,
                'EndDate': period_end,
                dimension_name: dimension_value,
                'Cost': amount,
                'Currency': currency,
                'Usage': usage
            })
    
    # Add the total cost for each period
    for time_period in cost_response['ResultsByTime']:
        period_start = time_period['TimePeriod']['Start']
        period_end = time_period['TimePeriod']['End']
        
        # For monthly granularity, use the month name
        date_obj = datetime.strptime(period_start, '%Y-%m-%d')
        if period_end > period_start:
            # This is for monthly granularity
            period_label = date_obj.strftime('%B %Y')
        else:
            # This is for daily granularity
            period_label = period_start
            
        if 'Total' in time_period:
            total_amount = float(time_period['Total']['UnblendedCost']['Amount'])
            currency = time_period['Total']['UnblendedCost']['Unit']
            total_usage = float(time_period['Total']['UsageQuantity']['Amount'])
            
            results.append({
                'Period': period_label,
                'StartDate': period_start,
                'EndDate': period_end,
                dimension_name: 'Total',
                'Cost': total_amount,
                'Currency': currency,
                'Usage': total_usage
            })
    
    return results

def create_excel_report(cost_data, output_file, quarter, year, group_by='service', granularity='monthly'):
    """Create an Excel report with the AWS cost data"""
    if not cost_data:
        print("No cost data available to create the report.")
        return False
    
    # Convert to DataFrame and ensure the dimension column has the right name
    df = pd.DataFrame(cost_data)
    dimension_name = group_by.capitalize()
    if group_by == 'tag':
        dimension_name = 'Tag'
    
    # Create a pivot table with dimensions as columns and periods as rows
    pivot_df = df.pivot_table(
        index='Period', 
        columns=dimension_name, 
        values='Cost', 
        aggfunc='sum'
    ).reset_index()
    
    # Sort the pivot table by start date to ensure chronological order
    period_to_date = {row['Period']: row['StartDate'] for _, row in df.drop_duplicates('Period').iterrows()}
    pivot_df['SortDate'] = pivot_df['Period'].map(period_to_date)
    pivot_df = pivot_df.sort_values('SortDate').drop('SortDate', axis=1)
    
    # Calculate quarterly total by dimension
    quarterly_totals = df.groupby(dimension_name)['Cost'].sum().reset_index()
    quarterly_totals = quarterly_totals.sort_values('Cost', ascending=False)
    
    # Create the Excel writer object
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Write the period breakdown DataFrame to Excel
    pivot_df.to_excel(writer, sheet_name=f"{granularity.capitalize()} Breakdown", index=False)
    
    # Write the quarterly totals to Excel
    quarterly_totals.to_excel(writer, sheet_name='Quarterly Summary', index=False)
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    breakdown_sheet = writer.sheets[f"{granularity.capitalize()} Breakdown"]
    summary_sheet = writer.sheets['Quarterly Summary']
    
    # Apply formatting (common style elements)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format Breakdown sheet
    # Format headers
    for col_num, column_title in enumerate(pivot_df.columns):
        cell = breakdown_sheet.cell(row=1, column=col_num+1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Set column width
        breakdown_sheet.column_dimensions[get_column_letter(col_num+1)].width = 15
    
    # Format all cells with borders and number format for cost columns
    for row_num in range(2, len(pivot_df) + 2):
        for col_num in range(1, len(pivot_df.columns) + 1):
            cell = breakdown_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Apply number format to cost columns (skip the Period column)
            if col_num > 1:
                cell.number_format = '$#,##0.00'
    
    # Add title and report info to the breakdown sheet
    breakdown_sheet.insert_rows(0, 2)
    breakdown_sheet['A1'] = f"AWS Q{quarter} {year} Cost Breakdown (by {granularity})"
    breakdown_sheet['A1'].font = Font(size=14, bold=True)
    breakdown_sheet.merge_cells('A1:E1')
    
    # Format Quarterly Summary sheet
    # Format headers
    for col_num, column_title in enumerate(quarterly_totals.columns):
        cell = summary_sheet.cell(row=1, column=col_num+1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Set column width
        summary_sheet.column_dimensions[get_column_letter(col_num+1)].width = 20
    
    # Format data cells
    for row_num in range(2, len(quarterly_totals) + 2):
        for col_num in range(1, len(quarterly_totals.columns) + 1):
            cell = summary_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Apply number format to cost column
            if col_num == 2:  # Cost column
                cell.number_format = '$#,##0.00'
    
    # Add percentage column to summary
    total_cost = quarterly_totals['Cost'].sum()
    summary_sheet['C1'] = 'Percentage'
    summary_sheet['C1'].fill = header_fill
    summary_sheet['C1'].font = header_font
    summary_sheet['C1'].alignment = Alignment(horizontal='center')
    summary_sheet['C1'].border = border
    
    for row_num in range(2, len(quarterly_totals) + 2):
        cost_cell = summary_sheet.cell(row=row_num, column=2)
        percentage = float(cost_cell.value) / total_cost if total_cost > 0 else 0
        
        cell = summary_sheet.cell(row=row_num, column=3)
        cell.value = percentage
        cell.number_format = '0.00%'
        cell.border = border
    
    # Add title and report info to the summary sheet
    summary_sheet.insert_rows(0, 2)
    summary_sheet['A1'] = f"AWS Q{quarter} {year} Cost Summary"
    summary_sheet['A1'].font = Font(size=14, bold=True)
    summary_sheet.merge_cells('A1:C1')
    
    # Add total row to summary
    total_row = len(quarterly_totals) + 3  # +3 because we added 2 title rows
    summary_sheet[f'A{total_row}'] = 'Total'
    summary_sheet[f'A{total_row}'].font = Font(bold=True)
    summary_sheet[f'B{total_row}'] = total_cost
    summary_sheet[f'B{total_row}'].number_format = '$#,##0.00'
    summary_sheet[f'B{total_row}'].font = Font(bold=True)
    summary_sheet[f'C{total_row}'] = 1
    summary_sheet[f'C{total_row}'].number_format = '0.00%'
    summary_sheet[f'C{total_row}'].font = Font(bold=True)
    
    # Create a pie chart for quarterly cost distribution
    pie = PieChart()
    pie.title = f"Q{quarter} {year} Cost Distribution by {dimension_name}"
    
    # Get data for top 10 items (to avoid cluttering the chart)
    max_items = min(10, len(quarterly_totals))
    data = Reference(summary_sheet, min_col=2, max_col=2, 
                     min_row=3, max_row=2+max_items)
    cats = Reference(summary_sheet, min_col=1, max_col=1, 
                     min_row=3, max_row=2+max_items)
    
    pie.add_data(data)
    pie.set_categories(cats)
    pie.height = 15
    pie.width = 10
    
    # Add the chart to the summary sheet
    summary_sheet.add_chart(pie, "E5")
    
    # Create a trend chart showing costs over the quarter
    chart_sheet = workbook.create_sheet('Cost Trends')
    
    if granularity == 'monthly':
        chart = BarChart()
    else:  # daily
        chart = LineChart()
    
    chart.title = f"AWS Cost Trends - Q{quarter} {year}"
    chart.x_axis.title = "Period"
    chart.y_axis.title = "Cost (USD)"
    
    # Get the column index for 'Total' if it exists
    if 'Total' in pivot_df.columns:
        total_col_idx = list(pivot_df.columns).index('Total') + 1
        
        # Add 2 to row references because of the title rows we inserted
        data = Reference(breakdown_sheet, min_col=total_col_idx, max_col=total_col_idx, 
                        min_row=3, max_row=len(pivot_df) + 3)
        cats = Reference(breakdown_sheet, min_col=1, max_col=1, 
                        min_row=4, max_row=len(pivot_df) + 3)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 15
        chart.width = 25
        
        chart_sheet.add_chart(chart, "B5")
    
    # Add title to chart sheet
    chart_sheet['A1'] = f"AWS Q{quarter} {year} Cost Trends"
    chart_sheet['A1'].font = Font(size=14, bold=True)
    
    # Create a comparison sheet for the top services/accounts/regions
    comparison_sheet = workbook.create_sheet('Top Cost Drivers')
    
    # Get the top 5 cost drivers
    top_five = quarterly_totals.head(5)
    
    # Generate a table showing the top 5 cost drivers
    comparison_sheet['A1'] = f"Top 5 Cost Drivers - Q{quarter} {year}"
    comparison_sheet['A1'].font = Font(size=14, bold=True)
    comparison_sheet.merge_cells('A1:C1')
    
    comparison_sheet['A3'] = dimension_name
    comparison_sheet['B3'] = 'Cost'
    comparison_sheet['C3'] = 'Percentage'
    
    for cell in ['A3', 'B3', 'C3']:
        comparison_sheet[cell].fill = header_fill
        comparison_sheet[cell].font = header_font
        comparison_sheet[cell].alignment = Alignment(horizontal='center')
        comparison_sheet[cell].border = border
        
    for idx, row in enumerate(top_five.itertuples(), start=4):
        comparison_sheet[f'A{idx}'] = row[1]  # Dimension value
        comparison_sheet[f'B{idx}'] = row[2]  # Cost
        comparison_sheet[f'B{idx}'].number_format = '$#,##0.00'
        
        percentage = row[2] / total_cost if total_cost > 0 else 0
        comparison_sheet[f'C{idx}'] = percentage
        comparison_sheet[f'C{idx}'].number_format = '0.00%'
        
        for col in ['A', 'B', 'C']:
            comparison_sheet[f'{col}{idx}'].border = border
    
    # Create a bar chart for the top 5
    top_chart = BarChart()
    top_chart.type = "col"
    top_chart.title = f"Top 5 {dimension_name} Cost Drivers"
    top_chart.x_axis.title = dimension_name
    top_chart.y_axis.title = "Cost (USD)"
    
    data = Reference(comparison_sheet, min_col=2, max_col=2, 
                     min_row=3, max_row=8)
    cats = Reference(comparison_sheet, min_col=1, max_col=1, 
                     min_row=4, max_row=8)
    
    top_chart.add_data(data, titles_from_data=True)
    top_chart.set_categories(cats)
    top_chart.height = 15
    top_chart.width = 20
    
    comparison_sheet.add_chart(top_chart, "E5")
    
    # Add report metadata
    info_sheet = workbook.create_sheet('Report Info')
    info_sheet['A1'] = "AWS Quarterly Cost Report - Metadata"
    info_sheet['A1'].font = Font(size=14, bold=True)
    
    info_sheet['A3'] = "Report Generated:"
    info_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    info_sheet['A4'] = "Report Period:"
    start_date, end_date = get_quarter_dates(year, quarter)
    info_sheet['B4'] = f"Q{quarter} {year} ({start_date} to {end_date})"
    
    info_sheet['A5'] = "Grouping Dimension:"
    info_sheet['B5'] = dimension_name
    
    info_sheet['A6'] = "Data Granularity:"
    info_sheet['B6'] = granularity.capitalize()
    
    info_sheet['A7'] = "Total Quarterly Cost:"
    info_sheet['B7'] = total_cost
    info_sheet['B7'].number_format = '$#,##0.00'
    
    info_sheet['A9'] = "Notes:"
    info_sheet['A10'] = "- Cost data is based on AWS Cost Explorer API and may have a delay of 24-48 hours"
    info_sheet['A11'] = "- The report includes unblended costs"
    info_sheet['A12'] = f"- Costs are grouped by {dimension_name.lower()}"
    
    # Set the first sheet as active
    workbook.active = 0
    
    # Save the workbook
    writer.close()
    
    return True

def main():
    # Parse command line arguments
    args = parse_arguments()
    
    print(f"Generating AWS cost report for Q{args.quarter} {args.year}...")
    
    # Get cost data
    cost_response = get_aws_quarterly_costs(
        args.year, 
        args.quarter, 
        group_by=args.group_by,
        tag_key=args.tag_key,
        profile=args.profile,
        region=args.region,
        granularity=args.granularity
    )
    
    if cost_response:
        # Parse cost data
        cost_data = parse_cost_data(cost_response, args.group_by)
        
        # Create Excel report
        if create_excel_report(cost_data, args.output, args.quarter, args.year, 
                               args.group_by, args.granularity):
            print(f"Excel report created successfully: {args.output}")
        else:
            print("Failed to create Excel report.")
    else:
        print("Failed to fetch AWS cost data.")

if __name__ == "__main__":
    main()
