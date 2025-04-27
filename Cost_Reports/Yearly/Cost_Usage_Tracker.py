import boto3
import pandas as pd
import calendar
from datetime import datetime, timedelta
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import argparse

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='Generate yearly AWS cost reports')
    parser.add_argument('--year', type=int, help='Year (YYYY)')
    parser.add_argument('--output', type=str, help='Output file path', default=None)
    parser.add_argument('--profile', type=str, help='AWS profile name', default=None)
    parser.add_argument('--region', type=str, help='AWS region', default='us-east-1')
    parser.add_argument('--group-by', type=str, choices=['service', 'account', 'region', 'tag'], 
                        default='service', help='Group costs by dimension')
    parser.add_argument('--tag-key', type=str, help='Tag key to group by (if group-by=tag)')
    parser.add_argument('--granularity', type=str, choices=['monthly', 'quarterly'], 
                        default='monthly', help='Data granularity')
    parser.add_argument('--top-n', type=int, default=10, 
                        help='Number of top items to highlight in charts and analysis')
    
    args = parser.parse_args()
    
    # Set default for year if not provided
    if args.year is None:
        # Default to previous year
        current_year = datetime.now().year
        args.year = current_year - 1
    
    # Set default output file if not provided
    if args.output is None:
        args.output = f"AWS_Yearly_Costs_{args.year}.xlsx"
    
    # Additional validation for tag-based grouping
    if args.group_by == 'tag' and args.tag_key is None:
        raise ValueError("--tag-key is required when --group-by=tag")
    
    return args

def get_year_dates(year):
    """Get start and end dates for a given year"""
    start_date = f"{year}-01-01"
    end_date = f"{year+1}-01-01"  # CE API end date is exclusive
    
    return start_date, end_date

def get_quarters_for_year(year):
    """Get start and end dates for each quarter in a year"""
    quarters = []
    
    for q in range(1, 5):
        if q == 1:
            start_month = 1
        elif q == 2:
            start_month = 4
        elif q == 3:
            start_month = 7
        elif q == 4:
            start_month = 10
        
        start_date = f"{year}-{start_month:02d}-01"
        
        if q < 4:
            end_month = start_month + 3
            end_year = year
        else:
            end_month = 1
            end_year = year + 1
        
        end_date = f"{end_year}-{end_month:02d}-01"
        quarter_name = f"Q{q} {year}"
        
        quarters.append({
            'name': quarter_name,
            'start_date': start_date,
            'end_date': end_date
        })
    
    return quarters

def get_aws_yearly_costs(year, group_by='service', tag_key=None, 
                         profile=None, region='us-east-1', granularity='monthly'):
    """Fetch yearly AWS cost data using Cost Explorer API"""
    try:
        # Get start and end dates for the year
        start_date, end_date = get_year_dates(year)
        
        # Create boto3 session with optional profile
        if profile:
            session = boto3.Session(profile_name=profile, region_name=region)
        else:
            session = boto3.Session(region_name=region)
        
        # Initialize the Cost Explorer client
        ce_client = session.client('ce')
        
        # Determine GroupBy configuration based on input
        if group_by == 'service':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'SERVICE'}]
        elif group_by == 'account':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'LINKED_ACCOUNT'}]
        elif group_by == 'region':
            group_by_config = [{'Type': 'DIMENSION', 'Key': 'REGION'}]
        elif group_by == 'tag' and tag_key:
            group_by_config = [{'Type': 'TAG', 'Key': tag_key}]
        else:
            raise ValueError(f"Invalid group_by: {group_by}")
        
        # Convert granularity string to API format
        ce_granularity = 'MONTHLY'
        if granularity == 'quarterly':
            # For quarterly, we'll fetch monthly and aggregate later
            ce_granularity = 'MONTHLY'
        
        # Request cost data from AWS Cost Explorer
        response = ce_client.get_cost_and_usage(
            TimePeriod={
                'Start': start_date,
                'End': end_date
            },
            Granularity=ce_granularity,
            Metrics=['UnblendedCost', 'UsageQuantity'],
            GroupBy=group_by_config
        )
        
        return response
    
    except Exception as e:
        print(f"Error fetching AWS cost data: {str(e)}")
        return None

def parse_cost_data(cost_response, group_by='service', year=None, granularity='monthly'):
    """Parse the AWS Cost Explorer response into a structured format"""
    if not cost_response:
        return None
    
    results = []
    
    # Determine the name of the dimension based on group_by
    dimension_name = group_by.capitalize()
    if group_by == 'tag':
        dimension_name = 'Tag'
    
    # Get quarters for the year (for quarterly aggregation)
    quarters = get_quarters_for_year(year) if year else []
    
    # Create a mapping of dates to quarters
    date_to_quarter = {}
    for quarter in quarters:
        start = quarter['start_date']
        end = quarter['end_date']
        current = start
        
        # For each month in this quarter, map to quarter name
        while current < end:
            date_to_quarter[current] = quarter['name']
            
            # Move to next month
            year_month = current.split('-')
            year_val = int(year_month[0])
            month_val = int(year_month[1])
            
            if month_val == 12:
                next_year = year_val + 1
                next_month = 1
            else:
                next_year = year_val
                next_month = month_val + 1
                
            current = f"{next_year}-{next_month:02d}-01"
    
    for time_period in cost_response['ResultsByTime']:
        period_start = time_period['TimePeriod']['Start']
        period_end = time_period['TimePeriod']['End']
        
        # For monthly granularity, use the month name
        date_obj = datetime.strptime(period_start, '%Y-%m-%d')
        
        # Set the period label based on granularity
        if granularity == 'monthly':
            period_label = date_obj.strftime('%B %Y')
        else:  # quarterly
            period_label = date_to_quarter.get(period_start, 'Unknown')
        
        for group in time_period['Groups']:
            dimension_value = group['Keys'][0]
            
            # For services, strip the prefix
            if group_by == 'service' and dimension_value.startswith('Amazon'):
                dimension_value = dimension_value.replace('Amazon ', '')
            
            amount = float(group['Metrics']['UnblendedCost']['Amount'])
            currency = group['Metrics']['UnblendedCost']['Unit']
            usage = float(group['Metrics']['UsageQuantity']['Amount'])
            
            results.append({
                'Period': period_label,
                'StartDate': period_start,
                'EndDate': period_end,
                dimension_name: dimension_value,
                'Cost': amount,
                'Currency': currency,
                'Usage': usage,
                'QuarterPeriod': date_to_quarter.get(period_start, 'Unknown')
            })
    
    # Add the total cost for each period
    for time_period in cost_response['ResultsByTime']:
        period_start = time_period['TimePeriod']['Start']
        period_end = time_period['TimePeriod']['End']
        
        # For monthly granularity, use the month name
        date_obj = datetime.strptime(period_start, '%Y-%m-%d')
        
        # Set the period label based on granularity
        if granularity == 'monthly':
            period_label = date_obj.strftime('%B %Y')
        else:  # quarterly
            period_label = date_to_quarter.get(period_start, 'Unknown')
            
        if 'Total' in time_period:
            total_amount = float(time_period['Total']['UnblendedCost']['Amount'])
            currency = time_period['Total']['UnblendedCost']['Unit']
            total_usage = float(time_period['Total']['UsageQuantity']['Amount'])
            
            results.append({
                'Period': period_label,
                'StartDate': period_start,
                'EndDate': period_end,
                dimension_name: 'Total',
                'Cost': total_amount,
                'Currency': currency,
                'Usage': total_usage,
                'QuarterPeriod': date_to_quarter.get(period_start, 'Unknown')
            })
    
    # For quarterly granularity, aggregate the monthly data
    if granularity == 'quarterly':
        # Convert to DataFrame for easier aggregation
        df = pd.DataFrame(results)
        
        # Group by quarter and dimension
        quarterly_df = df.groupby(['QuarterPeriod', dimension_name]).agg({
            'Cost': 'sum',
            'Usage': 'sum',
            'Currency': 'first'  # Assuming currency is the same for all entries
        }).reset_index()
        
        # Rename QuarterPeriod to Period
        quarterly_df = quarterly_df.rename(columns={'QuarterPeriod': 'Period'})
        
        # Convert back to list of dictionaries
        results = quarterly_df.to_dict('records')
    
    return results

def create_excel_report(cost_data, output_file, year, group_by='service', 
                        granularity='monthly', top_n=10):
    """Create an Excel report with the AWS cost data"""
    if not cost_data:
        print("No cost data available to create the report.")
        return False
    
    # Convert to DataFrame
    df = pd.DataFrame(cost_data)
    dimension_name = group_by.capitalize()
    if group_by == 'tag':
        dimension_name = 'Tag'
    
    # Create a pivot table with dimensions as columns and periods as rows
    pivot_df = df.pivot_table(
        index='Period', 
        columns=dimension_name, 
        values='Cost', 
        aggfunc='sum'
    ).reset_index()
    
    # Sort the pivot table chronologically
    month_order = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12,
        'Q1': 1, 'Q2': 2, 'Q3': 3, 'Q4': 4
    }
    
    def sort_key(period):
        if granularity == 'monthly':
            # Extract month name
            month_name = period.split()[0]
            return month_order.get(month_name, 0)
        elif granularity == 'quarterly':
            # Extract quarter
            quarter = period.split()[0]
            return month_order.get(quarter, 0)
        return 0
    
    # Sort by period
    pivot_df['sort_key'] = pivot_df['Period'].apply(sort_key)
    pivot_df = pivot_df.sort_values('sort_key').drop('sort_key', axis=1)
    
    # Calculate yearly total by dimension
    yearly_totals = df.groupby(dimension_name)['Cost'].sum().reset_index()
    yearly_totals = yearly_totals.sort_values('Cost', ascending=False)
    
    # Create quarter-by-quarter comparison if we have monthly data
    quarter_comparison = None
    if granularity == 'monthly':
        # Add quarter information if not already there
        if 'QuarterPeriod' not in df.columns:
            quarters = get_quarters_for_year(year)
            date_to_quarter = {}
            for quarter in quarters:
                start = quarter['start_date']
                end = quarter['end_date']
                current = start
                
                while current < end:
                    date_to_quarter[current] = quarter['name']
                    
                    # Move to next month
                    year_month = current.split('-')
                    year_val = int(year_month[0])
                    month_val = int(year_month[1])
                    
                    if month_val == 12:
                        next_year = year_val + 1
                        next_month = 1
                    else:
                        next_year = year_val
                        next_month = month_val + 1
                        
                    current = f"{next_year}-{next_month:02d}-01"
            
            df['QuarterPeriod'] = df['StartDate'].map(date_to_quarter)
        
        # Group by quarter and dimension
        quarter_comparison = df.groupby(['QuarterPeriod', dimension_name]).agg({
            'Cost': 'sum'
        }).reset_index()
        
        # Create a pivot table for quarter comparison
        quarter_pivot = quarter_comparison.pivot_table(
            index='QuarterPeriod',
            columns=dimension_name,
            values='Cost',
            aggfunc='sum'
        ).reset_index()
        
        # Sort by quarter
        def quarter_sort_key(quarter):
            if quarter.startswith('Q1'):
                return 1
            elif quarter.startswith('Q2'):
                return 2
            elif quarter.startswith('Q3'):
                return 3
            elif quarter.startswith('Q4'):
                return 4
            return 0
        
        quarter_pivot['sort_key'] = quarter_pivot['QuarterPeriod'].apply(quarter_sort_key)
        quarter_pivot = quarter_pivot.sort_values('sort_key').drop('sort_key', axis=1)
    
    # Create the Excel writer object
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Write the period breakdown DataFrame to Excel
    pivot_df.to_excel(writer, sheet_name=f"{granularity.capitalize()} Breakdown", index=False)
    
    # Write the yearly totals to Excel
    yearly_totals.to_excel(writer, sheet_name='Yearly Summary', index=False)
    
    # Write quarter comparison if available
    if quarter_comparison is not None and granularity == 'monthly':
        quarter_pivot.to_excel(writer, sheet_name='Quarterly Comparison', index=False)
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    breakdown_sheet = writer.sheets[f"{granularity.capitalize()} Breakdown"]
    summary_sheet = writer.sheets['Yearly Summary']
    quarter_sheet = None
    if quarter_comparison is not None and granularity == 'monthly':
        quarter_sheet = writer.sheets['Quarterly Comparison']
    
    # Apply formatting (common style elements)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format Breakdown sheet
    # Format headers
    for col_num, column_title in enumerate(pivot_df.columns):
        cell = breakdown_sheet.cell(row=1, column=col_num+1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Set column width
        breakdown_sheet.column_dimensions[get_column_letter(col_num+1)].width = 15
    
    # Format all cells with borders and number format for cost columns
    for row_num in range(2, len(pivot_df) + 2):
        for col_num in range(1, len(pivot_df.columns) + 1):
            cell = breakdown_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Apply number format to cost columns (skip the Period column)
            if col_num > 1:
                cell.number_format = '$#,##0.00'
    
    # Add title and report info to the breakdown sheet
    breakdown_sheet.insert_rows(0, 2)
    breakdown_sheet['A1'] = f"AWS {year} Cost Breakdown (by {granularity})"
    breakdown_sheet['A1'].font = Font(size=14, bold=True)
    breakdown_sheet.merge_cells('A1:E1')
    
    # Format Yearly Summary sheet
    # Format headers
    for col_num, column_title in enumerate(yearly_totals.columns):
        cell = summary_sheet.cell(row=1, column=col_num+1)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Set column width
        summary_sheet.column_dimensions[get_column_letter(col_num+1)].width = 20
    
    # Format data cells
    for row_num in range(2, len(yearly_totals) + 2):
        for col_num in range(1, len(yearly_totals.columns) + 1):
            cell = summary_sheet.cell(row=row_num, column=col_num)
            cell.border = border
            
            # Apply number format to cost column
            if col_num == 2:  # Cost column
                cell.number_format = '$#,##0.00'
    
    # Add percentage column to summary
    total_cost = yearly_totals['Cost'].sum()
    summary_sheet['C1'] = 'Percentage'
    summary_sheet['C1'].fill = header_fill
    summary_sheet['C1'].font = header_font
    summary_sheet['C1'].alignment = Alignment(horizontal='center')
    summary_sheet['C1'].border = border
    
    for row_num in range(2, len(yearly_totals) + 2):
        cost_cell = summary_sheet.cell(row=row_num, column=2)
        percentage = float(cost_cell.value) / total_cost if total_cost > 0 else 0
        
        cell = summary_sheet.cell(row=row_num, column=3)
        cell.value = percentage
        cell.number_format = '0.00%'
        cell.border = border
    
    # Add title and report info to the summary sheet
    summary_sheet.insert_rows(0, 2)
    summary_sheet['A1'] = f"AWS {year} Cost Summary"
    summary_sheet['A1'].font = Font(size=14, bold=True)
    summary_sheet.merge_cells('A1:C1')
    
    # Add total row to summary
    total_row = len(yearly_totals) + 3  # +3 because we added 2 title rows
    summary_sheet[f'A{total_row}'] = 'Total'
    summary_sheet[f'A{total_row}'].font = Font(bold=True)
    summary_sheet[f'B{total_row}'] = total_cost
    summary_sheet[f'B{total_row}'].number_format = '$#,##0.00'
    summary_sheet[f'B{total_row}'].font = Font(bold=True)
    summary_sheet[f'C{total_row}'] = 1
    summary_sheet[f'C{total_row}'].number_format = '0.00%'
    summary_sheet[f'C{total_row}'].font = Font(bold=True)
    
    # Format Quarterly Comparison sheet if it exists
    if quarter_sheet:
        # Format headers
        for col_num, column_title in enumerate(quarter_pivot.columns):
            cell = quarter_sheet.cell(row=1, column=col_num+1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
            # Set column width
            quarter_sheet.column_dimensions[get_column_letter(col_num+1)].width = 15
        
        # Format all cells with borders and number format for cost columns
        for row_num in range(2, len(quarter_pivot) + 2):
            for col_num in range(1, len(quarter_pivot.columns) + 1):
                cell = quarter_sheet.cell(row=row_num, column=col_num)
                cell.border = border
                
                # Apply number format to cost columns (skip the Quarter column)
                if col_num > 1:
                    cell.number_format = '$#,##0.00'
        
        # Add title
        quarter_sheet.insert_rows(0, 2)
        quarter_sheet['A1'] = f"AWS {year} Quarterly Cost Comparison"
        quarter_sheet['A1'].font = Font(size=14, bold=True)
        quarter_sheet.merge_cells('A1:E1')
    
    # Create a pie chart for yearly cost distribution
    pie = PieChart()
    pie.title = f"{year} Cost Distribution by {dimension_name}"
    
    # Get data for top N items (to avoid cluttering the chart)
    max_items = min(top_n, len(yearly_totals))
    data = Reference(summary_sheet, min_col=2, max_col=2, 
                     min_row=3, max_row=2+max_items)
    cats = Reference(summary_sheet, min_col=1, max_col=1, 
                     min_row=3, max_row=2+max_items)
    
    pie.add_data(data)
    pie.set_categories(cats)
    pie.height = 15
    pie.width = 10
    
    # Add the chart to the summary sheet
    summary_sheet.add_chart(pie, "E5")
    
    # Create a trend chart showing costs over the year
    trend_sheet = workbook.create_sheet('Cost Trends')
    
    if 'Total' in pivot_df.columns:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"AWS Cost Trends - {year}"
        chart.x_axis.title = "Period"
        chart.y_axis.title = "Cost (USD)"
        
        # Get the column index for 'Total'
        total_col_idx = list(pivot_df.columns).index('Total') + 1
        
        # Add 2 to row references because of the title rows we inserted
        data = Reference(breakdown_sheet, min_col=total_col_idx, max_col=total_col_idx, 
                        min_row=3, max_row=len(pivot_df) + 3)
        cats = Reference(breakdown_sheet, min_col=1, max_col=1, 
                        min_row=4, max_row=len(pivot_df) + 3)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 15
        chart.width = 25
        
        trend_sheet.add_chart(chart, "B5")
    
    # Add title to trend sheet
    trend_sheet['A1'] = f"AWS {year} Cost Trends"
    trend_sheet['A1'].font = Font(size=14, bold=True)
    
    # Create a quarter-over-quarter comparison chart if we have quarterly data
    if quarter_sheet:
        qoq_chart = BarChart()
        qoq_chart.type = "col"
        qoq_chart.title = f"Quarterly Cost Comparison - {year}"
        qoq_chart.x_axis.title = "Quarter"
        qoq_chart.y_axis.title = "Cost (USD)"
        
        if 'Total' in quarter_pivot.columns:
            # Get the column index for 'Total'
            q_total_col_idx = list(quarter_pivot.columns).index('Total') + 1
            
            # Add 2 to row references because of the title rows we inserted
            q_data = Reference(quarter_sheet, min_col=q_total_col_idx, max_col=q_total_col_idx, 
                              min_row=3, max_row=len(quarter_pivot) + 3)
            q_cats = Reference(quarter_sheet, min_col=1, max_col=1, 
                              min_row=4, max_row=len(quarter_pivot) + 3)
            
            qoq_chart.add_data(q_data, titles_from_data=True)
            qoq_chart.set_categories(q_cats)
            qoq_chart.height = 15
            qoq_chart.width = 25
            
            quarter_sheet.add_chart(qoq_chart, "B20")
    
    # Create a top services/accounts/regions analysis
    top_sheet = workbook.create_sheet('Top Cost Drivers')
    
    # Get the top N cost drivers
    top_n_items = yearly_totals.head(top_n)
    
    # Generate a table showing the top N cost drivers
    top_sheet['A1'] = f"Top {top_n} Cost Drivers - {year}"
    top_sheet['A1'].font = Font(size=14, bold=True)
    top_sheet.merge_cells('A1:C1')
    
    top_sheet['A3'] = dimension_name
    top_sheet['B3'] = 'Cost'
    top_sheet['C3'] = 'Percentage'
    
    for cell in ['A3', 'B3', 'C3']:
        top_sheet[cell].fill = header_fill
        top_sheet[cell].font = header_font
        top_sheet[cell].alignment = Alignment(horizontal='center')
        top_sheet[cell].border = border
        
    for idx, row in enumerate(top_n_items.itertuples(), start=4):
        top_sheet[f'A{idx}'] = row[1]  # Dimension value
        top_sheet[f'B{idx}'] = row[2]  # Cost
        top_sheet[f'B{idx}'].number_format = '$#,##0.00'
        
        percentage = row[2] / total_cost if total_cost > 0 else 0
        top_sheet[f'C{idx}'] = percentage
        top_sheet[f'C{idx}'].number_format = '0.00%'
        
        for col in ['A', 'B', 'C']:
            top_sheet[f'{col}{idx}'].border = border
    
    # Create a bar chart for the top N
    top_chart = BarChart()
    top_chart.type = "col"
    top_chart.title = f"Top {top_n} {dimension_name} Cost Drivers - {year}"
    top_chart.x_axis.title = dimension_name
    top_chart.y_axis.title = "Cost (USD)"
    
    data = Reference(top_sheet, min_col=2, max_col=2, 
                     min_row=3, max_row=3+top_n)
    cats = Reference(top_sheet, min_col=1, max_col=1, 
                     min_row=4, max_row=3+top_n)
    
    top_chart.add_data(data, titles_from_data=True)
    top_chart.set_categories(cats)
    top_chart.height = 15
    top_chart.width = 20
    
    top_sheet.add_chart(top_chart, "E5")
    
    # Create a YoY comparison placeholder sheet
    yoy_sheet = workbook.create_sheet('YoY Comparison')
    yoy_sheet['A1'] = "Year-over-Year Cost Comparison"
    yoy_sheet['A1'].font = Font(size=14, bold=True)
    
    yoy_sheet['A3'] = "To track year-over-year changes, run this report for multiple years and consolidate the data here."
    yoy_sheet['A5'] = "Year"
    yoy_sheet['B5'] = "Total Cost"
    yoy_sheet['C5'] = "YoY Change (%)"
    
    for cell in ['A5', 'B5', 'C5']:
        yoy_sheet[cell].fill = header_fill
        yoy_sheet[cell].font = header_font
        yoy_sheet[cell].alignment = Alignment(horizontal='center')
        yoy_sheet[cell].border = border
    
    # Add example rows
    yoy_sheet['A6'] = year - 1
    yoy_sheet['B6'] = ""
    yoy_sheet['C6'] = ""
    
    yoy_sheet['A7'] = year
    yoy_sheet['B7'] = total_cost
    yoy_sheet['B7'].number_format = '$#,##0.00'
    yoy_sheet['C7'] = ""
    
    for row in range(6, 8):
        for col in ['A', 'B', 'C']:
            yoy_sheet[f'{col}{row}'].border = border
    
    # Add report metadata
    info_sheet = workbook.create_sheet('Report Info')
    info_sheet['A1'] = "AWS Yearly Cost Report - Metadata"
    info_sheet['A1'].font = Font(size=14, bold=True)
    
    info_sheet['A3'] = "Report Generated:"
    info_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    info_sheet['A4'] = "Report Period:"
    start_date, end_date = get_year_dates(year)
    info_sheet['B4'] = f"FY {year} ({start_date} to {end_date})"
    
    info_sheet['A5'] = "Grouping Dimension:"
    info_sheet['B5'] = dimension_name
    
    info_sheet['A6'] = "Data Granularity:"
    info_sheet['B6'] = granularity.capitalize()
    
    info_sheet['A7'] = "Total Yearly Cost:"
    info_sheet['B7'] = total_cost
    info_sheet['B7'].number_format = '$#,##0.00'
    
    info_sheet['A9'] = "Notes:"
    info_sheet['A10'] = "- Cost data is based on AWS Cost Explorer API and may have a delay of 24-48 hours"
    info_sheet['A11'] = "- The report includes unblended costs"
    info_sheet['A12'] = f"- Costs are grouped by {dimension_name.lower()}"
    info_sheet['A13'] = "- For more detailed analysis, consider exporting the data to a BI tool"
    
    # Set the first sheet as active
    workbook.active = 0
    
    # Save the workbook
    writer.close()
    
    return True

def main():
    # Parse command line arguments
    args = parse_arguments()
    
    print(f"Generating AWS cost report for {args.year}...")
    
    # Get cost data
    cost_response = get_aws_yearly_costs(
        args.year, 
        group_by=args.group_by,
        tag_key=args.tag_key,
        profile=args.profile,
        region=args.region,
        granularity=args.granularity
    )
    
    if cost_response:
        # Parse cost data
        cost_data = parse_cost_data(
            cost_response, 
            args.group_by, 
            args.year, 
            args.granularity
        )
        
        # Create Excel report
        if create_excel_report(
            cost_data, 
            args.output, 
            args.year, 
            args.group_by, 
            args.granularity,
            args.top_n
        ):
            print(f"Excel report created successfully: {args.output}")
        else:
            print("Failed to create Excel report.")
    else:
        print("Failed to fetch AWS cost data.")

if __name__ == "__main__":
    main()
